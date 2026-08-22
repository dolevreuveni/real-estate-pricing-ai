"""Tests for pricing_evidence."""
import re
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.data.pricing_evidence import (
    PRICING_EVIDENCE_COLUMNS,
    generate_run_id,
    save_pricing_evidence_csv,
    save_pricing_evidence_excel,
    validate_evidence,
)


def _sample_evidence() -> pd.DataFrame:
    rows = [
        {
            "run_id": "pricing_run_20260822_153000",
            "target_apartment_id": 1,
            "evidence_type": "transaction",
            "address": "Helsinki 20",
            "record_date": "2024-01-15",
            "rooms": 3,
            "area_sqm": 70.0,
            "floor": 2,
            "original_price": 4_000_000,
            "adjusted_price": 4_200_000,
            "price_per_sqm": 60_000,
            "distance_from_project_km": 0.2,
            "similarity_score": None,
            "pricing_weight": None,
            "source": "tax_authority",
            "source_url": None,
            "data_retrieved_at": "2026-08-01",
        },
        {
            "run_id": "pricing_run_20260822_153000",
            "target_apartment_id": 1,
            "evidence_type": "market_listing",
            "address": "Helsinki 22",
            "record_date": "2026-07-01",
            "rooms": 3,
            "area_sqm": 69.0,
            "floor": 3,
            "original_price": None,
            "adjusted_price": None,
            "price_per_sqm": 62_000,
            "distance_from_project_km": 0.1,
            "similarity_score": None,
            "pricing_weight": None,
            "source": "yad2",
            "source_url": None,
            "data_retrieved_at": "2026-08-01",
        },
        {
            "run_id": "pricing_run_20260822_153000",
            "target_apartment_id": 1,
            "evidence_type": "competitor_project",
            "address": "Arlozorov 5",
            "record_date": None,
            "rooms": 3,
            "area_sqm": 71.0,
            "floor": None,
            "original_price": None,
            "adjusted_price": None,
            "price_per_sqm": 63_000,
            "distance_from_project_km": 0.5,
            "similarity_score": None,
            "pricing_weight": None,
            "source": "madlan",
            "source_url": None,
            "data_retrieved_at": "2026-08-01",
        },
    ]
    return pd.DataFrame(rows, columns=PRICING_EVIDENCE_COLUMNS)


def test_valid_evidence_dataframe_passes_validation():
    validate_evidence(_sample_evidence())  # should not raise


def test_missing_required_column_raises():
    evidence = _sample_evidence().drop(columns=["source"])
    with pytest.raises(ValueError, match="missing required column"):
        validate_evidence(evidence)


def test_unknown_evidence_type_raises():
    evidence = _sample_evidence()
    evidence.loc[0, "evidence_type"] = "something_else"
    with pytest.raises(ValueError, match="unknown evidence_type"):
        validate_evidence(evidence)


def test_saving_csv(tmp_path: Path):
    evidence = _sample_evidence()
    path = tmp_path / "pricing_evidence.csv"
    result_path = save_pricing_evidence_csv(evidence, path)

    assert result_path == path
    assert path.exists()

    reloaded = pd.read_csv(path)
    assert list(reloaded.columns) == PRICING_EVIDENCE_COLUMNS
    assert len(reloaded) == 3


def test_saving_excel(tmp_path: Path):
    evidence = _sample_evidence()
    path = tmp_path / "pricing_evidence.xlsx"
    result_path = save_pricing_evidence_excel(evidence, path)

    assert result_path == path
    assert path.exists()


def test_excel_sheets_are_split_by_evidence_type(tmp_path: Path):
    evidence = _sample_evidence()
    path = tmp_path / "pricing_evidence.xlsx"
    save_pricing_evidence_excel(evidence, path)

    workbook = load_workbook(path)
    assert set(workbook.sheetnames) == {
        "all_evidence",
        "transactions_used",
        "market_listings_used",
        "competitor_projects_used",
    }

    all_evidence_df = pd.read_excel(path, sheet_name="all_evidence")
    transactions_df = pd.read_excel(path, sheet_name="transactions_used")
    listings_df = pd.read_excel(path, sheet_name="market_listings_used")
    competitors_df = pd.read_excel(path, sheet_name="competitor_projects_used")

    assert len(all_evidence_df) == 3
    assert len(transactions_df) == 1
    assert transactions_df.iloc[0]["evidence_type"] == "transaction"
    assert len(listings_df) == 1
    assert listings_df.iloc[0]["evidence_type"] == "market_listing"
    assert len(competitors_df) == 1
    assert competitors_df.iloc[0]["evidence_type"] == "competitor_project"


def test_generate_run_id_format():
    fixed_moment = datetime(2026, 8, 22, 15, 30, 0, tzinfo=timezone.utc)
    run_id = generate_run_id(fixed_moment)
    assert run_id == "pricing_run_20260822_153000"


def test_generate_run_id_default_matches_expected_pattern():
    run_id = generate_run_id()
    assert re.fullmatch(r"pricing_run_\d{8}_\d{6}", run_id)
