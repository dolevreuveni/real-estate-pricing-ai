"""One-time, reproducible enrichment of the raw apartment source Excel
(data/raw/Apartment_example.xlsx) with new SYNTHETIC POC attribute
columns that were not supplied by the company.

Added source columns (Hebrew headers, matching the file's existing style):
    מספר חניות      (parking_count)
    שטח מחסן        (storage_area_sqm)
    כיוון מרפסת     (balcony_direction)
    שטח גינה        (garden_area_sqm)
    שטח גג מוצמד    (roof_area_sqm)
    קומה אחרונה     (is_top_floor)

Every existing cell (apartment number, rooms, floor, apartment area,
balcony area, directions, notes) is left completely untouched -- this
script only appends new columns (H-M) to each row.

SYNTHETIC vs DERIVED -- documented per field:

* parking_count, storage_area_sqm, garden_area_sqm, roof_area_sqm,
  balcony_direction are SYNTHETIC POC assumptions, not supplied by the
  company. Each uses a small, deterministic, documented rule (never
  random) so the result is reproducible and auditable:

    parking_count:
        triplex/duplex (מקבל the "טריפלקס"/"דופלקס" note) -> 2
        rooms == 5                                        -> 2
        rooms == 2                                        -> 0
        otherwise (rooms == 3, the common case)            -> 1

    storage_area_sqm:
        triplex/duplex -> 8.0 sqm
        rooms == 5     -> 6.0 sqm
        rooms == 2     -> 0.0 sqm (no storage)
        otherwise      -> 4.0 sqm

    balcony_direction:
        a deterministic 8-point compass cycle keyed by apartment
        number: (apartment_number - 1) % 8 -> one of
        [East, South-East, South, South-West, West, North-West,
         North, North-East] (Hebrew terms). Deliberately independent of
        the existing "כיווני אוויר" (directions) column -- balcony
        direction and general apartment air-direction are different
        concepts (see src/data/apartment_reader.py).

    garden_area_sqm:
        garden apartments (note "דירת גן") -> 50% of that apartment's
        own interior area, rounded to 1 decimal. All other apartments -> 0.

    roof_area_sqm:
        triplex/duplex apartments -> 30% of the apartment's TOTAL
        interior area (summed across all of its source rows -- the
        roof is one shared amenity for the whole multi-level unit, not
        one per floor), rounded to 1 decimal. All other apartments -> 0.
        The same computed total is repeated on every source row
        belonging to that apartment (never divided/split across rows),
        so that normalization can safely aggregate with max() rather
        than sum() without double-counting.

* is_top_floor is DERIVED, not synthetic: a row's value is True only
  when that row's own floor number equals the maximum floor number
  found anywhere in the building (11, reached only by the triplex
  apartments 36/37 at their top level). This is computed directly from
  the real floor data already in the sheet. Apartment-level is_top_floor
  (post-normalization) is the boolean OR (max()) across an apartment's
  rows -- an apartment counts as "top floor" if ANY of its levels
  reaches the building's highest floor.

Run:
    python scripts/enrich_apartment_raw_excel.py
"""
import sys
from pathlib import Path

if __name__ == "__main__" and __package__ is None:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from src.config.settings import RAW_DATA_DIR

RAW_PATH = RAW_DATA_DIR / "Apartment_example.xlsx"

GROUND_FLOOR_LABEL = "קרקע"
TOTAL_ROW_LABEL = 'סה"כ'
GARDEN_LABEL = "דירת גן"
DUPLEX_LABEL = "דופלקס"
TRIPLEX_LABEL = "טריפלקס"

NEW_HEADERS = [
    "מספר חניות",
    "שטח מחסן",
    "כיוון מרפסת",
    "שטח גינה",
    "שטח גג מוצמד",
    "קומה אחרונה",
]

BALCONY_DIRECTION_CYCLE = [
    "מזרח",
    "דרום-מזרח",
    "דרום",
    "דרום-מערב",
    "מערב",
    "צפון-מערב",
    "צפון",
    "צפון-מזרח",
]


def _floor_number(value):
    if isinstance(value, str) and value.strip() == GROUND_FLOOR_LABEL:
        return 0
    if value is None:
        return None
    return int(value)


def main() -> None:
    wb = openpyxl.load_workbook(RAW_PATH)
    ws = wb.active

    header_row = 1
    for offset, header in enumerate(NEW_HEADERS):
        ws.cell(row=header_row, column=7 + 1 + offset, value=header)

    # First pass: read every data row and compute each apartment's total
    # interior area (needed for garden_area_sqm / roof_area_sqm, which
    # are apartment-level, not per-row-level, values).
    rows = []
    max_floor = 0
    apartment_total_area = {}
    for row_idx in range(2, ws.max_row + 1):
        floor_raw = ws.cell(row=row_idx, column=1).value
        apartment_number = ws.cell(row=row_idx, column=2).value
        rooms = ws.cell(row=row_idx, column=3).value
        area = ws.cell(row=row_idx, column=4).value
        notes = ws.cell(row=row_idx, column=7).value or ""

        is_total_row = isinstance(floor_raw, str) and floor_raw.strip() == TOTAL_ROW_LABEL
        rows.append((row_idx, floor_raw, apartment_number, rooms, area, notes, is_total_row))

        if not is_total_row:
            floor_num = _floor_number(floor_raw)
            if floor_num is not None:
                max_floor = max(max_floor, floor_num)
            apartment_total_area[apartment_number] = (
                apartment_total_area.get(apartment_number, 0.0) + float(area)
            )

    # Second pass: write the new columns.
    for row_idx, floor_raw, apartment_number, rooms, area, notes, is_total_row in rows:
        if is_total_row:
            # Summary rows are excluded during normalization regardless
            # -- leave the new synthetic columns blank rather than
            # implying they carry meaningful data.
            continue

        is_garden = GARDEN_LABEL in notes
        is_multilevel = (DUPLEX_LABEL in notes) or (TRIPLEX_LABEL in notes)
        floor_num = _floor_number(floor_raw)

        if is_multilevel:
            parking_count = 2
            storage_area_sqm = 8.0
        elif rooms == 5:
            parking_count = 2
            storage_area_sqm = 6.0
        elif rooms == 2:
            parking_count = 0
            storage_area_sqm = 0.0
        else:
            parking_count = 1
            storage_area_sqm = 4.0

        balcony_direction = BALCONY_DIRECTION_CYCLE[(int(apartment_number) - 1) % 8]

        garden_area_sqm = round(float(area) * 0.5, 1) if is_garden else 0.0
        roof_area_sqm = (
            round(apartment_total_area[apartment_number] * 0.3, 1) if is_multilevel else 0.0
        )

        is_top_floor = floor_num is not None and floor_num == max_floor

        ws.cell(row=row_idx, column=8, value=parking_count)
        ws.cell(row=row_idx, column=9, value=storage_area_sqm)
        ws.cell(row=row_idx, column=10, value=balcony_direction)
        ws.cell(row=row_idx, column=11, value=garden_area_sqm)
        ws.cell(row=row_idx, column=12, value=roof_area_sqm)
        ws.cell(row=row_idx, column=13, value=is_top_floor)

    wb.save(RAW_PATH)
    print(f"Enriched {RAW_PATH} with columns: {NEW_HEADERS}")
    print(f"Building max floor (used for is_top_floor derivation): {max_floor}")


if __name__ == "__main__":
    main()
